#coding=utf-8

import requests
from phone import Phone
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
"""
htmldown 此函数功能是 输入url连接，然后保存到本地进行备份。多用在有翻页的后台进行后台备份，输入翻页url，参数等进行批量down备份
url: 访问url
filename: 保存网页信息的本地目录
head: 请求header

"""



def htmldown(url,filename,head):
	for i in range(0,4):
		try:
			res=requests.get(url,header=head,timeout=5)
			with open(filename,'a+') as f:
				f.write(res.text)
			print(f"{url} successful!!!")
			return 
		except Exception as e:
			print(e)
			pass


"""
htmldown_post 此函数功能是 输入url连接，然后保存到本地进行备份。多用在有翻页的后台进行后台备份，输入翻页url，参数等进行批量down备份
[请求方式为post请求]
参数：
url: 访问url
filename: 保存网页信息的本地目录
head: 请求header
data1: post请求的参数
"""
def htmldown_post(url,filename,head,data1):
	for i in range(0,4):
		try:

			res=requests.post(url,header=head,data=data1,timeout=5)
			with open(filename,'a+') as f:
				f.write(res.text)
			print(f"{url}+'\t'+{data1} successful!!!")
			return 
		except Exception as e:
			print(e)
			pass

"""
ls_header 此函数提供对burp抓包的关键信息提取，可以提取url，提取其中的headers信息

"""

# head="""
# GET /userInfo/getUserBasicInfo HTTP/1.1
# Host: admin.ceshi123.com
# Connection: close
# Accept: */*
# User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36
# X-Requested-With: XMLHttpRequest
# Accept-Encoding: gzip, deflate
# Accept-Language: zh-CN,zh;q=0.9,en;q=0.8
# Cookie: phpsession=3478173981273981ceshi
# """


def ls_header(str_headers: str) -> dict:
    if not str_headers:
        return {}
    items = str_headers.splitlines()
    headers = {}
    for item in items:
        item_str = item.strip()
        if not item_str:
            continue
        if item_str.startswith(':'):
            continue
        i = item_str.find(':')
        headers[item_str[:i]] = item_str[i + 1:].strip()
    return headers



"""
phone_city 手机号归属地批量检测生成
输入手机号列表，返回 call+省+市 的列表

"""

def phone_city(calls):
	try:
		call = Phone()
		a=[]
		for i in calls:
			try:
				call1=call.find(i.strip())
				pro=call1['province']
				ci=call1['city']
				info=i+'\t'+pro+'\t'+ci
				# print(info)
				a.append(info)
			except Exception as e:
				# print(e)
				pass
		return a
	except Exception as e:
		print(e)
		pass


"""
read_xlsx()读取xlsx文件 结果返回列表

"""

def read_xlsx(f):
	try:
		file1=f
		xls_list=[]
		table = load_workbook(f)
		sheetname=table.sheetnames
		for shee in sheetname:
			data=table[shee]

			rows=data.rows
			for i in rows:
				a=[]
				for x in i:
					# print(x.value)
					a.append(str.lower(str(x.value)))
				xls_list.append(a)

		return xls_list
	except Exception as e:
		print(e)
		pass

"""
write_xlsx 此函数用来生成xlsx文件
传入参数必须是列表，多个函数组合调用来生成excel文件
"""

def write_xlsx(datas,fi):
	try:
		wb=Workbook()
		ws=wb.active

		for i in datas:
			ws.append(i)
		wb.save(fi)
		print(f"save file as {fi}!!! ")
	except Exception as e:
		print(e)


"""
multi_thread 多线程函数
single_thread: 要执行的函数
args: 传入的参数
number: 要执行的多线程数


"""


def multi_thread(single_thread,args,number):
    thread_pool = ThreadPoolExecutor(max_workers=number)
    r = [thread_pool.submit(single_thread,(*arg)) for arg in args]
    return [i.result() for i in r]


"""
后续讲持续更新模块，主渗透测试后的数据处理，渗透测测试模块
"""

